"""A synthetic workbook shaped exactly like a real ST export, so the round
trip can be tested without the network or the shipped files."""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from stproducts.api import parse_grid


def _col(**kw):
    base = {"id": "1", "name": "X", "order": "1", "show": True, "type": "string",
            "identifier": "I"}
    base.update(kw)
    return base


GRID_PAYLOAD = {
    "levelTitle": "Test series",
    "columns": [
        _col(id="1", name="Part Number", order="1"),
        _col(id="2", name="Marketing Status", order="2"),
        _col(id="3", name="Flash Size", symbol="kB", conditional="Prog",
             type="float", order="3"),
        _col(id="4", name="Number of A/D Converters", qualifier="typ", type="integer",
             order="4", aggregation="A/D Converters 12-bit"),
        _col(id="5", name="Number of Channels", qualifier="typ", type="integer",
             order="5", aggregation="A/D Converters 12-bit"),
        # Present in the API, absent from the workbook -> must be appended.
        _col(id="6", name="FPU", type="multi", order="6"),
        _col(id="7", name="Dual-bank Flash", type="boolean", order="7"),
    ],
    "rows": [
        {"productId": "PF1", "cells": [
            {"columnId": "1", "value": "TEST001"},
            {"columnId": "2", "value": "Active"},
            {"columnId": "3", "value": "128"},
            {"columnId": "4", "value": "3"},
            {"columnId": "5", "value": "16"},
            {"columnId": "6", "value": "Yes"},
            {"columnId": "7", "value": "false"},
        ]},
        {"productId": "PF2", "cells": [
            {"columnId": "1", "value": "TEST002"},
            {"columnId": "2", "value": "Active"},
            {"columnId": "3", "value": "256"},
            {"columnId": "4", "value": "3"},
            {"columnId": "5", "value": "16"},
        ]},
        # In ST's data, absent from the workbook -> NEW_PART.
        {"productId": "PF3", "cells": [
            {"columnId": "1", "value": "TEST003"},
            {"columnId": "2", "value": "Preview"},
            {"columnId": "3", "value": "512"},
        ]},
    ],
}


@pytest.fixture
def grid():
    return parse_grid("SS0001", GRID_PAYLOAD)


@pytest.fixture
def workbook_path(tmp_path):
    """Five columns, a grouped pair, ST's banner/breadcrumb/header skeleton.

    Seeded with three deliberate faults: a wrong Flash Size, a typo in
    Marketing Status, and a part ST no longer lists.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ProductsList"

    headers = [
        ("Part Number", None),
        ("Marketing Status", None),
        ("Flash Size (kB) (Prog)", None),
        ("A/D Converters 12-bit", "Number of A/D Converters typ"),
        ("A/D Converters 12-bit", "Number of Channels typ"),
    ]
    last = get_column_letter(len(headers))
    worksheet.merge_cells(f"A1:{last}8")
    worksheet.merge_cells(f"A9:{last}9")
    worksheet.cell(9, 1, "Microcontrollers & microprocessors/Test series")

    worksheet.merge_cells(start_row=10, start_column=4, end_row=10, end_column=5)
    worksheet.cell(10, 4, "A/D Converters 12-bit")
    for index, (head, sub) in enumerate(headers, start=1):
        if sub:
            worksheet.cell(11, index, sub)
        else:
            worksheet.merge_cells(start_row=10, start_column=index,
                                  end_row=11, end_column=index)
            worksheet.cell(10, index, head)

    rows = [
        ["TEST001", "Active", "999", "3", "16"],      # Flash Size is wrong
        ["TEST002", "Activ", "256.0", "3", "16"],     # status typo; 256.0 == 256
        ["TEST999", "Active", "64", "1", "8"],        # ST does not list this part
    ]
    for offset, row in enumerate(rows):
        for index, value in enumerate(row, start=1):
            worksheet.cell(12 + offset, index, value)

    path = tmp_path / "Test series - Products.xlsx"
    workbook.save(path)
    return path
