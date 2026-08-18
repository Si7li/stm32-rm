"""Reading a workbook, diffing it, and writing the corrected pair."""

from __future__ import annotations

import hashlib

from openpyxl import load_workbook

from stproducts.diffing import (
    ADDED_COLUMN,
    CHANGED,
    NEW_PART,
    NOT_IN_ST_DATA,
    compare,
    plan_columns,
)
from stproducts.compose import api_only_sheet
from stproducts.sheetio import read_original
from stproducts.writer import write_corrected, write_diff


def _composed(sheet, grid):
    keys = sheet.column_keys + [
        c.key for c in grid.columns if c.key not in sheet.column_keys
    ]
    return api_only_sheet(grid, keys)


def test_reads_the_export_skeleton(workbook_path):
    sheet = read_original(workbook_path)
    assert sheet.stem == "Test series"
    assert sheet.has_sub_header is True
    assert sheet.data_start_row == 12
    assert sheet.level_title == "Test series"
    assert sheet.parts == ["TEST001", "TEST002", "TEST999"]
    assert sheet.column_keys == [
        "Part Number",
        "Marketing Status",
        "Flash Size (kB) (Prog)",
        "A/D Converters 12-bit | Number of A/D Converters typ",
        "A/D Converters 12-bit | Number of Channels typ",
    ]


def test_grouped_header_continuation_is_not_read_as_empty(workbook_path):
    """The merged cell under a group title reads as None from openpyxl; the
    reader has to fill it in or the column loses its identity."""
    sheet = read_original(workbook_path)
    assert sheet.columns[4].group == "A/D Converters 12-bit"
    assert sheet.columns[4].label == "Number of Channels typ"


def _ready(workbook_path, grid):
    sheet = read_original(workbook_path)
    return sheet, grid, _composed(sheet, grid)


def test_diff_classifies_each_kind(workbook_path, grid):
    result = compare(*_ready(workbook_path, grid))
    kinds = {(r.part, r.column): r for r in result.records}

    changed = kinds[("TEST001", "Flash Size (kB) (Prog)")]
    assert changed.kind == CHANGED and changed.old == "999" and changed.new == "128"

    assert kinds[("TEST002", "Marketing Status")].kind == CHANGED

    # 256 vs 256.0 is formatting, not an error.
    assert ("TEST002", "Flash Size (kB) (Prog)") not in kinds

    assert kinds[("TEST003", "*")].kind == NEW_PART
    assert kinds[("TEST999", "*")].kind == NOT_IN_ST_DATA
    assert kinds[("TEST001", "FPU")].kind == ADDED_COLUMN

    assert result.counts[CHANGED] == 2
    assert result.parts_compared == 2
    assert result.missing_parts == ["TEST999"]
    assert result.new_parts == ["TEST003"]


def test_empty_appended_column_is_not_reported(workbook_path, grid):
    """TEST002 has no FPU value, so there is nothing to add for it."""
    result = compare(*_ready(workbook_path, grid))
    assert ("TEST002", "FPU") not in {(r.part, r.column) for r in result.records}


def test_column_plan_keeps_order_then_appends(workbook_path, grid):
    sheet = read_original(workbook_path)
    original, appended = plan_columns(sheet, grid)
    assert [k for k, _ in original] == sheet.column_keys
    assert [k for k, _ in appended] == ["FPU", "Dual-bank Flash"]


def test_corrected_workbook_layout(tmp_path, workbook_path, grid):
    sheet = read_original(workbook_path)
    out = tmp_path / "out.xlsx"
    write_corrected(out, sheet, grid, _composed(sheet, grid))

    rebuilt = read_original(out)
    # Original columns first, in order; then the API's extras; then the link.
    assert rebuilt.column_keys[: len(sheet.column_keys)] == sheet.column_keys
    assert rebuilt.column_keys[len(sheet.column_keys):] == [
        "FPU", "Dual-bank Flash", "Datasheet URL",
    ]
    # Every original part survives, and ST's new one is there too.
    assert set(sheet.parts) <= set(rebuilt.parts)
    assert "TEST003" in rebuilt.parts
    # Original row order is preserved; ST's new part is appended at the end.
    assert rebuilt.parts == ["TEST001", "TEST002", "TEST003", "TEST999"]
    # Corrected values, in ST's own rendering.
    assert rebuilt.data["TEST001"]["Flash Size (kB) (Prog)"] == "128"
    assert rebuilt.data["TEST001"]["Dual-bank Flash"] == "No"
    assert rebuilt.data["TEST002"]["Marketing Status"] == "Active"
    assert rebuilt.data["TEST003"]["A/D Converters 12-bit | Number of Channels typ"] == "-"


def test_part_kept_when_st_does_not_list_it(tmp_path, workbook_path, grid):
    sheet = read_original(workbook_path)
    out = tmp_path / "out.xlsx"
    write_corrected(out, sheet, grid, _composed(sheet, grid))
    rebuilt = read_original(out)
    assert "TEST999" in rebuilt.parts
    assert rebuilt.data["TEST999"]["Datasheet URL"] == NOT_IN_ST_DATA
    assert rebuilt.data["TEST999"]["Flash Size (kB) (Prog)"] == "64"  # not overwritten


def test_diff_workbook_lists_every_record(tmp_path, workbook_path, grid):
    result = compare(*_ready(workbook_path, grid))
    out = tmp_path / "diff.xlsx"
    write_diff(out, result, level_id="SS0001", level_title="Test series")

    worksheet = load_workbook(out).active
    head = next(r for r in range(1, 40) if worksheet.cell(r, 1).value == "Part Number")
    written = worksheet.max_row - head
    assert written == len(result.records)
    assert worksheet.cell(head, 5).value == "Class"


def test_output_is_byte_identical_across_writes(tmp_path, workbook_path, grid):
    """Validation item 10 in miniature: the bytes depend on content only."""
    sheet = read_original(workbook_path)
    digests = []
    for name in ("a.xlsx", "b.xlsx"):
        path = tmp_path / name
        write_corrected(path, sheet, grid, _composed(sheet, grid))
        digests.append(hashlib.sha256(path.read_bytes()).hexdigest())
    assert digests[0] == digests[1]
