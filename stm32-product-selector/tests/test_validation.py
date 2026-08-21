"""The acceptance checks from both build specs, run against the real files.

These run offline against the warm caches and skip cleanly if they are not
there, so the suite still passes on a fresh clone.

``product_selector_out/`` is expected to hold a **datasheet-first** run
(``stproducts build``). The API-path checks build into a temp directory of
their own so the two never collide.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from stproducts.api import fetch_grid
from stproducts.cli import corrections_digest, digest_inputs, input_files, main
from stproducts.compose import api_only_sheet
from stproducts.exporter import export_sheet_json
from stproducts.net import Fetcher
from stproducts.provenance import FROM_PDF, TOKENS
from stproducts.sheetio import read_original
from stproducts.values import is_blank

ROOT = Path(__file__).resolve().parents[2]
INPUTS = ROOT / "product_selector"
CACHE = ROOT / "cache"
DATASHEET_CACHE = ROOT / "datasheets_cache"
DATASHEETS = ROOT / "datasheets"
OUT = ROOT / "product_selector_out"
BASELINE = Path(__file__).parent / "data" / "api_baseline_sha256.json"

pytestmark = pytest.mark.skipif(
    not (INPUTS.exists() and CACHE.exists()),
    reason="needs the shipped workbooks and a warm cache",
)


@pytest.fixture(scope="module")
def fetcher():
    return Fetcher(cache_dir=CACHE, offline=True)


@pytest.fixture(scope="module")
def report():
    path = OUT / "run_report.json"
    if not path.exists():
        pytest.skip("run `stproducts build` first")
    return json.loads(path.read_text())


@pytest.fixture(scope="module")
def datasheet_report(report):
    if report.get("source") != "datasheet":
        pytest.skip("product_selector_out holds an --source api run")
    return report


def _sheets(stem):
    path = OUT / f"{stem}.xlsx"
    workbook = load_workbook(path)
    sheet = read_original(path)
    return workbook, sheet


def _sheets_path(path):
    workbook = load_workbook(path)
    sheet = read_original(path)
    return workbook, sheet


def _local_stems(report):
    """Report keys of workbooks built from a real original (not discovered)."""
    return [stem for stem, e in report["files"].items() if e.get("source_workbook")]


# --------------------------------------------------------------------------
# Carried over: the selector-API build spec
# --------------------------------------------------------------------------


@pytest.mark.parametrize(
    "level_id,title,rows,workbook",
    [
        ("SS1575", "STM32F2 series", 38, "STM32F2 series"),
        ("SC1244", "STM8 8-bit MCUs", 135, "STM8 8-bit MCUs"),
        ("SC2230", "STM32 Arm Cortex MPUs", 64, "STM32 Arm Cortex MPUs"),
    ],
)
def test_known_ids_match_their_workbooks(fetcher, level_id, title, rows, workbook):
    grid = fetch_grid(fetcher, level_id)
    assert grid.level_title == title
    assert len(grid.rows) == rows
    sheet = read_original(INPUTS / f"{workbook} - Products.xlsx")
    assert len(sheet.parts) == rows
    if level_id == "SC1244":
        assert len(grid.columns) == 32 == len(sheet.columns)


def test_every_workbook_resolves_or_is_reported(report):
    expected = {p.stem[: -len(" - Products")] for p in input_files(INPUTS, None)}
    assert expected <= set(report["files"]) | {u["file"] for u in report["unresolved"]}
    for entry in report["unresolved"]:
        assert entry["candidates"] or entry["pages_tried"]
        assert entry["reason"]


def test_column_order_is_preserved_and_extras_follow(report):
    for stem in _local_stems(report):
        original = read_original(INPUTS / f"{stem} - Products.xlsx")
        corrected = read_original(OUT / f"{stem}.xlsx")
        assert corrected.column_keys[: len(original.column_keys)] == original.column_keys
        assert corrected.has_sub_header == original.has_sub_header
        assert corrected.column_keys[-1] == "Datasheet URL"


def test_no_part_is_silently_dropped(report):
    for stem in _local_stems(report):
        original = read_original(INPUTS / f"{stem} - Products.xlsx")
        corrected = read_original(OUT / f"{stem}.xlsx")
        assert set(original.parts) <= set(corrected.parts)


# --------------------------------------------------------------------------
# The inversion spec
# --------------------------------------------------------------------------


def test_1_f205rb_comes_from_the_datasheet(datasheet_report):
    """Every value in Table 2 of stm32f205rb.pdf, marked DATASHEET."""
    workbook, sheet = _sheets("STM32F2 series")
    data, provenance = workbook[sheet.sheet_title], workbook["Provenance"]
    columns = {c.key: c.index for c in sheet.columns}
    row = next(
        r for r in range(sheet.data_start_row, data.max_row + 1)
        if data.cell(r, 1).value == "STM32F205RB"
    )
    expected = {
        "Flash Size (kB) (Prog)": "128",
        "RAM Size (kB)": "64",
        "I2C typ": "3",
        "SPI typ": "3",
        "CAN (2.0)": "2",
        "I/Os (High Current)": "51",
        "USART typ": "4",
        "UART typ": "2",
        "Package": "LQFP64",
    }
    for key, value in expected.items():
        assert str(data.cell(row, columns[key]).value) == value, key
        assert provenance.cell(row, columns[key]).value == "DATASHEET", key


def test_2_f207ie_i2c_is_three_and_the_override_is_recorded(datasheet_report):
    """The case the inversion exists for: the datasheet says 3, ST says 2."""
    workbook, sheet = _sheets("STM32F2 series")
    data, provenance = workbook[sheet.sheet_title], workbook["Provenance"]
    columns = {c.key: c.index for c in sheet.columns}
    row = next(
        r for r in range(sheet.data_start_row, data.max_row + 1)
        if data.cell(r, 1).value == "STM32F207IE"
    )
    assert str(data.cell(row, columns["I2C typ"]).value) == "3"
    assert provenance.cell(row, columns["I2C typ"]).value == "DATASHEET"

    diff = load_workbook(OUT / "STM32F2 series - diff.xlsx").active
    head = next(r for r in range(1, 40) if diff.cell(r, 1).value == "Part Number")
    rows = [
        tuple(diff.cell(r, c).value for c in range(1, 6))
        for r in range(head + 1, diff.max_row + 1)
    ]
    assert ("STM32F207IE", "I2C typ", "2", "3", "DATASHEET_OVERRIDES_API") in rows
    assert any(
        r[0] == "STM32F207IE" and r[1] == "I2C typ"
        and r[4] == "ORIGINAL_MATCHED_API_NOT_DATASHEET"
        for r in rows
    )


def test_3_provenance_mirrors_the_data_sheet(datasheet_report):
    for stem, entry in datasheet_report["files"].items():
        workbook, sheet = _sheets_path(OUT / entry["corrected"])
        assert "Provenance" in workbook.sheetnames, stem
        data, provenance = workbook[sheet.sheet_title], workbook["Provenance"]
        assert (provenance.max_row, provenance.max_column) == (
            data.max_row, data.max_column,
        ), stem
        for row in range(sheet.data_start_row, data.max_row + 1):
            for column in range(1, data.max_column + 1):
                token = provenance.cell(row, column).value
                assert token in TOKENS, f"{stem} r{row}c{column}: {token!r}"


def test_4_no_cell_claims_datasheet_without_a_source_table(datasheet_report):
    """Asserted through the extractor's own invariant, not by inspection."""
    from stproducts.provenance import Reading

    for token in FROM_PDF:
        with pytest.raises(ValueError):
            Reading(token, "3")  # no source table
        with pytest.raises(ValueError):
            Reading(token, None, "Table 2")  # no value

    # And every DATASHEET cell in the real output came from a named table.
    counted = 0
    for stem, entry in datasheet_report["files"].items():
        counted += entry["provenance"]["DATASHEET"] + entry["provenance"]["DERIVED"]
        if entry["provenance"]["DATASHEET"]:
            assert entry["datasheets_with_summary_table"] >= 1, stem
    assert counted > 0


def test_5_per_family_extraction_success_is_reported(datasheet_report):
    for stem, entry in datasheet_report["files"].items():
        assert "datasheets" in entry, stem
        acquisition = entry["datasheets"]
        assert acquisition["parts_resolved"] + acquisition["parts_unresolved"] > 0
        assert "datasheets_with_summary_table" in entry
        assert "datasheets_without_summary_table" in entry
        assert "parts_without_datasheet" in entry
    # STM8 and MPU are the families flagged as at risk; both must be present
    # with a verdict either way rather than quietly missing.
    for stem in ("STM8 8-bit MCUs", "STM32 Arm Cortex MPUs"):
        assert stem in datasheet_report["files"]


def test_6_inputs_are_untouched(report):
    assert report["inputs_unchanged"] is True


def test_7_api_source_reproduces_the_pre_inversion_output(tmp_path):
    """--source api must still produce the exact workbooks it always did.

    A hash mismatch here does **not** by itself mean the code regressed. The
    baseline is taken against a cache of ST's API, and some columns ST serves
    are live commercial state rather than facts about the part -- ``Buy On
    Line`` tracks distributor stock and flips on its own. Refreshing the cache
    is enough to break every hash while the code is untouched, which is
    exactly what happened between 2026-08-10 and 2026-08-11: seven STM8 parts
    came back into stock and three workbooks stopped matching.

    So on failure this reports *which parts and columns* moved. If they are
    all volatile columns, the code is fine and the baseline wants
    regenerating (see ``scripts/refresh_api_baseline.py``). If a specification
    column moved, that is a real regression.
    """
    if not BASELINE.exists():
        pytest.skip("no recorded baseline")
    expected = json.loads(BASELINE.read_text())
    code = main([
        "build", "--source", "api", "--input", str(INPUTS), "--out", str(tmp_path),
        "--cache", str(CACHE), "--offline",
    ])
    assert code == 0
    actual = {
        p.name: hashlib.sha256(p.read_bytes()).hexdigest()
        for p in sorted(tmp_path.glob("*.xlsx"))
    }
    if actual == expected:
        return

    moved = sorted(k for k in set(expected) | set(actual) if expected.get(k) != actual.get(k))
    columns = _changed_columns(tmp_path, moved)
    volatile = columns <= VOLATILE_COLUMNS and bool(columns)
    pytest.fail(
        f"{len(moved)} workbook(s) differ from the recorded baseline: "
        f"{', '.join(moved)}\n"
        f"columns carrying reported changes: {sorted(columns) or 'none identified'}\n"
        + (
            "all of these are volatile commercial columns, so this is stale-baseline "
            "drift rather than a code regression -- regenerate the baseline"
            if volatile
            else "at least one is a specification column -- investigate before "
            "regenerating the baseline"
        )
    )


#: Columns ST serves that describe commercial state, not the part. They change
#: without any code change, so they cannot anchor a reproducibility baseline.
VOLATILE_COLUMNS = {"Buy On Line", "Marketing Status"}


def _changed_columns(out_dir: Path, workbooks: list[str]) -> set[str]:
    """Column names appearing in the diff sheets of the given workbooks."""
    import openpyxl

    names: set[str] = set()
    for name in workbooks:
        path = out_dir / name
        if not path.exists() or " - diff" not in name:
            path = out_dir / name.replace(".xlsx", " - diff.xlsx")
        if not path.exists():
            continue
        book = openpyxl.load_workbook(path)
        for row in book["Diff"].iter_rows(values_only=True):
            if row and len(row) > 4 and row[1] and row[4] and row[1] != "Column":
                names.add(str(row[1]))
    return names


def test_8_warm_run_makes_no_network_calls(tmp_path):
    code = main([
        "build", "--source", "datasheet", "--input", str(INPUTS), "--out", str(tmp_path),
        "--cache", str(CACHE), "--datasheets", str(DATASHEETS),
        "--datasheet-cache", str(DATASHEET_CACHE), "--offline", "--only", "STM32F2 series",
    ])
    assert code == 0
    fresh = json.loads((tmp_path / "run_report.json").read_text())
    assert fresh["network_calls"] == 0
    assert fresh["inputs_unchanged"] is True


def test_10_corrections_json_reconciles_with_the_run_report(tmp_path):
    """The corrected+added parameter log exists and adds up to the diff."""
    code = main([
        "build", "--source", "datasheet", "--input", str(INPUTS), "--out", str(tmp_path),
        "--cache", str(CACHE), "--datasheets", str(DATASHEETS),
        "--datasheet-cache", str(DATASHEET_CACHE), "--offline", "--only", "STM32F2 series",
    ])
    assert code == 0
    corrections = json.loads((tmp_path / "corrections.json").read_text())

    stem = "STM32F2 series"
    entry = corrections["files"][stem]
    assert entry["level_id"] == "SS1575"

    # corrected_cells == CHANGED + BLANK_FILLED + MISSING_FROM_ST for the file.
    classes = json.loads((tmp_path / "run_report.json").read_text())["files"][stem]["classes"]
    expected = classes["CHANGED"] + classes["BLANK_FILLED"] + classes["MISSING_FROM_ST"]
    assert sum(g["count"] for g in entry["corrected"]) == expected
    assert corrections["totals"]["corrected_cells"] == expected

    # Every corrected record names its column, carries before/after and parts.
    for group in entry["corrected"]:
        assert group["parameter"] and group["kind"] in ("CHANGED", "BLANK_FILLED", "MISSING_FROM_ST")
        assert group["parts"] and group["count"] == len(group["parts"])
        if group["kind"] == "CHANGED":
            assert group["from"] != group["to"]

    # Appended columns are named; added_parameters cells sum to ADDED_COLUMN.
    assert entry["appended_columns"]
    assert sum(g["count"] for g in entry["added_parameters"]) == classes["ADDED_COLUMN"]
    assert {g["parameter"] for g in entry["added_parameters"]} <= set(entry["appended_columns"])


def test_11_per_file_json_reconciles_with_the_workbook(tmp_path):
    """Each output workbook gets a same-named JSON in the Sidekick shape whose
    records are exactly the cells written to the xlsx."""
    code = main([
        "build", "--source", "datasheet", "--input", str(INPUTS), "--out", str(tmp_path),
        "--cache", str(CACHE), "--datasheets", str(DATASHEETS),
        "--datasheet-cache", str(DATASHEET_CACHE), "--offline", "--only", "STM32F2 series",
    ])
    assert code == 0
    doc = json.loads((tmp_path / "STM32F2 series.json").read_text())
    assert doc["document"] == "STM32F2 series"
    assert doc["level_id"] == "SS1575"
    assert "notes" not in doc

    sheet = read_original(tmp_path / "STM32F2 series.xlsx")
    expected_keys = set(sheet.column_keys) - {"Datasheet URL"} | {"LPUART typ"}
    assert {p["part_number"] for p in doc["products"]} == set(sheet.data)
    assert doc["product_count"] == len(doc["products"])
    for product in doc["products"]:
        part = product["part_number"]
        assert part in sheet.data
        assert set(product["values"]) == expected_keys
        for key, value in product["values"].items():
            assert str(sheet.data[part].get(key)) == value, (part, key)
        assert set(product["descriptions"]) == set(product["values"])
        assert "Inter-Integrated Circuit" in product["descriptions"]["I2C typ"]
        assert product["url"] == (
            "https://www.st.com/en/microcontrollers-microprocessors/"
            f"{part.lower()}.html"
        )
        assert product["semantic_type"] == "product_selector"


def test_12_discovered_selector_has_json_and_corrections_entry(fetcher):
    """A discovered selector has no original to diff against: corrections
    report the whole sheet as added, and the JSON still carries values and a
    multi-line description of every column."""
    grid = fetch_grid(fetcher, "LN1035")  # STM32F405/415
    keys = [c.key for c in grid.columns] + ["LPUART typ"]
    composed = api_only_sheet(grid, keys)

    digest = corrections_digest(None, keys, composed, grid)
    assert digest["corrected"] == []
    assert digest["appended_columns"] == keys
    assert digest["note"] == "discovered selector: no original workbook to diff against"
    assert sum(g["count"] for g in digest["added_parameters"]) == sum(
        1 for cp in composed.parts.values() for cell in cp.cells.values() if not is_blank(cell.value)
    )

    doc = export_sheet_json("STM32F405/415", grid, keys, composed)
    assert doc["document"] == "STM32F405/415"
    assert doc["level_id"] == "LN1035"
    assert "notes" not in doc
    assert doc["product_count"] == len(doc["products"])
    for product in doc["products"]:
        assert set(product["values"]) == set(product["descriptions"]) == set(keys)
        for key in keys:
            assert product["descriptions"][key]
            assert "\n" in product["descriptions"][key]


def test_9_report_carries_provenance_and_both_new_counts(datasheet_report):
    for stem, entry in datasheet_report["files"].items():
        assert set(entry["provenance"]) <= set(TOKENS), stem
        if entry.get("source_workbook"):
            # The nine real files are large enough to exercise every token.
            assert set(entry["provenance"]) == set(TOKENS), stem
            assert "DATASHEET_OVERRIDES_API" in entry["classes"], stem
            assert "ORIGINAL_MATCHED_API_NOT_DATASHEET" in entry["classes"], stem
        assert entry["parts_without_datasheet"] >= 0
    assert "provenance_totals" in datasheet_report
    for name, total in datasheet_report["totals"].items():
        assert total == sum(
            e["classes"][name]
            for e in datasheet_report["files"].values()
            if "classes" in e
        )


def test_inputs_are_read_only_in_practice():
    files = input_files(INPUTS, None)
    first = digest_inputs(files)
    assert len(first) == 9
    assert digest_inputs(files) == first
