"""Write the corrected workbook and the diff workbook.

The corrected file reproduces the original's skeleton -- banner and logo,
breadcrumb, header row, and the sub-header row for files that group columns
-- keeps every original column in its original position, and appends the
columns the API carries that the original lacked, in the API's own order.

Output is deterministic: no timestamps, no run-varying metadata, so a second
run over a warm cache is byte-identical (validation item 10).
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .api import Column, Grid
from .compose import ComposedSheet
from .diffing import NOT_IN_ST_DATA, DiffResult, plan_columns
from .provenance import AMBIGUOUS, API, DATASHEET, DERIVED, UNAVAILABLE
from .sheetio import BANNER_ROWS, BREADCRUMB_ROW, HEADER_ROW, OriginalSheet
from .values import render

HEADER_FILL = PatternFill("solid", fgColor="00C0C0C0")
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="thin"))
LINK_FONT = Font(name="Calibri", size=11, color="000000FF")
BODY_FONT = Font(name="Calibri", size=11)
FLAG_FILL = PatternFill("solid", fgColor="00FFF2CC")

#: Constant so two runs over the same data produce identical bytes.
FIXED_TIMESTAMP = datetime(2000, 1, 1, tzinfo=timezone.utc).replace(tzinfo=None)

DEFAULT_WIDTH = 18.0
PRODUCT_URL = "https://www.st.com/en/product/{part}.html"

#: Appended last so it never disturbs positional reads of the API columns.
DATASHEET_COLUMN = "Datasheet URL"

TOKEN_FILL = {
    DATASHEET: PatternFill("solid", fgColor="00C6EFCE"),
    DERIVED: PatternFill("solid", fgColor="00D9EAD3"),
    AMBIGUOUS: PatternFill("solid", fgColor="00FFEB9C"),
    API: PatternFill("solid", fgColor="00DDEBF7"),
    UNAVAILABLE: PatternFill("solid", fgColor="00F2F2F2"),
}


#: xlsx is a zip, and a zip stores a modification time per entry. openpyxl
#: stamps those with the clock, so two identical runs differ in bytes even
#: though every cell matches. Rewriting the archive with one fixed timestamp
#: is what actually makes validation item 10 hold.
ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)


#: openpyxl re-stamps dcterms:modified with the clock as it saves, ignoring
#: whatever the properties object said, so it is pinned here instead.
_MODIFIED = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")
FIXED_XML_TIMESTAMP = b"2000-01-01T00:00:00Z"


def _normalise_archive(path: Path) -> None:
    """Rewrite an xlsx so its bytes depend only on its content."""
    with zipfile.ZipFile(path) as source:
        entries = [(info, source.read(info.filename)) for info in source.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for info, payload in entries:
            if info.filename == "docProps/core.xml":
                payload = _MODIFIED.sub(rb"\g<1>" + FIXED_XML_TIMESTAMP + rb"\g<2>", payload)
            fresh = zipfile.ZipInfo(info.filename, date_time=ZIP_TIMESTAMP)
            fresh.compress_type = info.compress_type
            fresh.external_attr = info.external_attr
            fresh.internal_attr = info.internal_attr
            fresh.create_system = info.create_system
            target.writestr(fresh, payload)


def _save(workbook: Workbook, path: Path) -> None:
    _strip_deterministic(workbook)
    workbook.save(path)
    _normalise_archive(Path(path))


def _strip_deterministic(workbook: Workbook) -> None:
    """Pin the metadata that would otherwise vary between runs.

    openpyxl insists on writing dcterms:created/modified, so they are fixed
    to a constant rather than cleared -- the point is only that two runs
    produce the same bytes.
    """
    props = workbook.properties
    props.creator = "stproducts"
    props.lastModifiedBy = "stproducts"
    props.created = FIXED_TIMESTAMP
    props.modified = FIXED_TIMESTAMP
    props.title = None
    props.revision = None


def write_corrected(
    path: Path,
    sheet: OriginalSheet,
    grid: Grid,
    composed: ComposedSheet,
    *,
    datasheet_urls: dict[str, str] | None = None,
    provenance: bool = False,
    extras: bool = True,
) -> None:
    original_cols, appended_cols = plan_columns(sheet, grid, extras=extras)
    layout: list[tuple[str, Column | None]] = [*original_cols, *appended_cols]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet.sheet_title

    total = len(layout) + 1  # + Datasheet URL
    last_letter = get_column_letter(total)

    # Banner: same merged block and the original's own logo.
    worksheet.merge_cells(f"A1:{last_letter}{BANNER_ROWS}")
    if sheet.logo_png:
        image = XLImage(io.BytesIO(sheet.logo_png))
        if sheet.logo_size:
            image.width, image.height = sheet.logo_size
        worksheet.add_image(image, "A1")

    worksheet.merge_cells(f"A{BREADCRUMB_ROW}:{last_letter}{BREADCRUMB_ROW}")
    crumb = worksheet.cell(BREADCRUMB_ROW, 1, sheet.breadcrumb)
    crumb.font = Font(name="Calibri", size=11, bold=True)
    crumb.alignment = Alignment(horizontal="left", vertical="bottom")

    sub_row = HEADER_ROW + 1
    data_start = sub_row + 1 if sheet.has_sub_header else HEADER_ROW + 1

    # Header, with grouped columns merged across and labelled beneath.
    index = 0
    while index < len(layout):
        key, column = layout[index]
        group = column.aggregation if column is not None else _group_of(sheet, key)
        col_index = index + 1
        if group and sheet.has_sub_header:
            span = index
            while span < len(layout) and _group_for(sheet, layout[span]) == group:
                span += 1
            worksheet.merge_cells(
                start_row=HEADER_ROW, start_column=col_index,
                end_row=HEADER_ROW, end_column=span,
            )
            worksheet.cell(HEADER_ROW, col_index, group)
            for offset in range(index, span):
                label = layout[offset][0].split(" | ", 1)[-1]
                worksheet.cell(sub_row, offset + 1, label)
            index = span
            continue

        label = key.split(" | ", 1)[-1] if group else key
        if sheet.has_sub_header:
            worksheet.merge_cells(
                start_row=HEADER_ROW, start_column=col_index,
                end_row=sub_row, end_column=col_index,
            )
        worksheet.cell(HEADER_ROW, col_index, label)
        index += 1

    datasheet_col = len(layout) + 1
    if sheet.has_sub_header:
        worksheet.merge_cells(
            start_row=HEADER_ROW, start_column=datasheet_col,
            end_row=sub_row, end_column=datasheet_col,
        )
    worksheet.cell(HEADER_ROW, datasheet_col, DATASHEET_COLUMN)

    header_rows = (HEADER_ROW, sub_row) if sheet.has_sub_header else (HEADER_ROW,)
    for row in header_rows:
        for col in range(1, total + 1):
            cell = worksheet.cell(row, col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = HEADER_BORDER

    widths = {c.key: c.width for c in sheet.columns}
    for position, (key, _) in enumerate(layout, start=1):
        worksheet.column_dimensions[get_column_letter(position)].width = (
            widths.get(key) or DEFAULT_WIDTH
        )
    worksheet.column_dimensions[get_column_letter(datasheet_col)].width = 52.0

    # Body: the original workbook's row order is the order ST shipped it in,
    # and the file we ship replaces that one. Parts ST still lists keep their
    # original position; parts ST added since are genuinely new rows and go
    # at the end in API order; parts the original had that ST no longer lists
    # are carried over verbatim after them, flagged.
    api_rows = grid.rows_by_part()
    known = set(sheet.parts)
    ordered = [p for p in sheet.parts if p in api_rows]
    ordered += [p for p in grid.part_numbers if p not in known]
    kept = [p for p in sheet.parts if p not in api_rows]

    tokens: list[list[str]] = []
    conditions: list[tuple[str, str, str, str]] = []

    row_index = data_start
    for part in ordered:
        cells = composed.parts[part].cells
        row_tokens = []
        for position, (key, column) in enumerate(layout, start=1):
            cell = cells[key]
            worksheet.cell(row_index, position, cell.value).font = BODY_FONT
            row_tokens.append(cell.token)
            if cell.token == AMBIGUOUS and cell.conditions:
                conditions.append((part, key, cell.token, cell.conditions))
        url = (datasheet_urls or {}).get(part)
        worksheet.cell(row_index, datasheet_col, url or "-").font = BODY_FONT
        row_tokens.append(API if url else UNAVAILABLE)
        _link_part(worksheet.cell(row_index, 1), part)
        tokens.append(row_tokens)
        row_index += 1

    for part in kept:
        for position, (key, _) in enumerate(layout, start=1):
            value = sheet.data[part].get(key)
            cell = worksheet.cell(row_index, position, "-" if value is None else str(value).strip())
            cell.font = BODY_FONT
            cell.fill = FLAG_FILL
        flag = worksheet.cell(row_index, datasheet_col, NOT_IN_ST_DATA)
        flag.font = BODY_FONT
        flag.fill = FLAG_FILL
        _link_part(worksheet.cell(row_index, 1), part)
        # Neither source lists this part, so neither sourced its values: the
        # row is carried over verbatim from the original workbook.
        tokens.append([UNAVAILABLE] * (len(layout) + 1))
        row_index += 1

    worksheet.freeze_panes = worksheet.cell(data_start, 2)

    if provenance:
        _write_provenance(workbook, worksheet, sheet, layout, tokens, data_start)
        _write_conditions(workbook, conditions)

    _save(workbook, path)


def _write_provenance(workbook, data_sheet, sheet, layout, tokens, data_start) -> None:
    """A token grid the same shape as the data sheet.

    Kept as a separate sheet so the data sheet stays clean and usable, and
    dimension-for-dimension identical so a reader can line the two up cell by
    cell without counting.
    """
    provenance = workbook.create_sheet("Provenance")
    total = len(layout) + 1
    last = get_column_letter(total)

    provenance.merge_cells(f"A1:{last}{BANNER_ROWS}")
    provenance.cell(1, 1, f"Provenance for {sheet.stem} — one token per cell")
    provenance.merge_cells(f"A{BREADCRUMB_ROW}:{last}{BREADCRUMB_ROW}")
    provenance.cell(BREADCRUMB_ROW, 1, sheet.breadcrumb).font = Font(
        name="Calibri", size=11, bold=True
    )

    sub_row = HEADER_ROW + 1
    header_rows = (HEADER_ROW, sub_row) if sheet.has_sub_header else (HEADER_ROW,)
    for row in header_rows:
        for col in range(1, total + 1):
            source = data_sheet.cell(row, col)
            cell = provenance.cell(row, col, source.value)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = HEADER_BORDER

    for offset, row_tokens in enumerate(tokens):
        for position, token in enumerate(row_tokens, start=1):
            cell = provenance.cell(data_start + offset, position, token)
            cell.font = BODY_FONT
            fill = TOKEN_FILL.get(token)
            if fill:
                cell.fill = fill

    for position in range(1, total + 1):
        provenance.column_dimensions[get_column_letter(position)].width = 16
    provenance.freeze_panes = provenance.cell(data_start, 2)


def _write_conditions(workbook, conditions) -> None:
    """What the datasheet does say where it cannot settle a value."""
    sheet = workbook.create_sheet("Conditions")
    headers = ["Part Number", "Column", "Provenance", "Datasheet evidence"]
    for position, name in enumerate(headers, start=1):
        cell = sheet.cell(1, position, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = HEADER_ALIGN
    for offset, (part, column, token, text) in enumerate(conditions, start=2):
        sheet.cell(offset, 1, part)
        sheet.cell(offset, 2, column)
        sheet.cell(offset, 3, token)
        sheet.cell(offset, 4, text)
    for letter, width in zip("ABCD", (18, 44, 16, 110)):
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = sheet.cell(2, 1)


def _group_for(sheet: OriginalSheet, entry: tuple[str, Column | None]) -> str | None:
    key, column = entry
    if column is not None:
        return column.aggregation
    return _group_of(sheet, key)


def _group_of(sheet: OriginalSheet, key: str) -> str | None:
    for column in sheet.columns:
        if column.key == key:
            return column.group
    return key.split(" | ", 1)[0] if " | " in key else None


def _link_part(cell, part: str) -> None:
    cell.hyperlink = PRODUCT_URL.format(part=part.lower())
    cell.font = LINK_FONT


DIFF_HEADERS = ["Part Number", "Column", "Old value (workbook)", "New value (ST)", "Class"]

CLASS_FILL = {
    "CHANGED": PatternFill("solid", fgColor="00FFC7CE"),
    "BLANK_FILLED": PatternFill("solid", fgColor="00FFEB9C"),
    "MISSING_FROM_ST": PatternFill("solid", fgColor="00F2F2F2"),
    "ADDED_COLUMN": PatternFill("solid", fgColor="00DDEBF7"),
    "NEW_PART": PatternFill("solid", fgColor="00C6EFCE"),
    "NOT_IN_ST_DATA": PatternFill("solid", fgColor="00FFF2CC"),
    "DATASHEET_OVERRIDES_API": PatternFill("solid", fgColor="00B7E1CD"),
    "ORIGINAL_MATCHED_API_NOT_DATASHEET": PatternFill("solid", fgColor="00F4CCCC"),
}


def write_diff(path: Path, result: DiffResult, *, level_id: str, level_title: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Diff"

    summary = result.summary()
    lines = [
        ("Workbook", result.stem),
        ("ST level id", level_id),
        ("ST level title", level_title),
        ("Parts compared", summary["parts_compared"]),
        ("Cells compared", summary["cells_compared"]),
        ("Original columns", summary["original_columns"]),
        ("Appended columns", summary["appended_columns"]),
    ]
    for name, count in summary["classes"].items():
        lines.append((name, count))

    for offset, (name, value) in enumerate(lines, start=1):
        label = worksheet.cell(offset, 1, name)
        label.font = Font(name="Calibri", size=11, bold=True)
        worksheet.cell(offset, 2, value)

    head = len(lines) + 2
    for position, name in enumerate(DIFF_HEADERS, start=1):
        cell = worksheet.cell(head, position, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = HEADER_ALIGN

    for offset, record in enumerate(result.records, start=head + 1):
        worksheet.cell(offset, 1, record.part)
        worksheet.cell(offset, 2, record.column)
        worksheet.cell(offset, 3, record.old)
        worksheet.cell(offset, 4, record.new)
        cell = worksheet.cell(offset, 5, record.kind)
        fill = CLASS_FILL.get(record.kind)
        if fill:
            cell.fill = fill

    for letter, width in zip("ABCDE", (18, 44, 46, 46, 18)):
        worksheet.column_dimensions[letter].width = width
    worksheet.freeze_panes = worksheet.cell(head + 1, 1)
    worksheet.auto_filter.ref = (
        f"A{head}:E{head + len(result.records)}" if result.records else f"A{head}:E{head}"
    )

    _save(workbook, path)
