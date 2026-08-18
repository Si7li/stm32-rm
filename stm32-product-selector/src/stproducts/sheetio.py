"""Read an original product-selector workbook and recover its structure.

Every one of the nine exports has the same skeleton:

======  ==========================================================
rows    content
1-8     ST banner -- one merged block holding the ST logo image
9       breadcrumb, merged across the full width, bold
10      header row (grey fill, centred, thin bottom rule)
11      sub-header row -- only in files that use grouped columns,
        where row 10 carries a horizontally merged group title
        (``A/D Converters 12-bit``) over row 11's ``Number of
        Channels typ``. Files without groups start data at row 11.
12/11+  one row per part; part number is a blue hyperlink to
        ``/en/product/<part>.html``; absent values are written ``-``
======  ==========================================================

A grouped column's identity is ``"<group> | <sub-header>"``, which is
exactly :attr:`stproducts.api.Column.key`, so workbook columns and API
columns line up on a single string.
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BANNER_ROWS = 8
BREADCRUMB_ROW = 9
HEADER_ROW = 10


@dataclass
class SheetColumn:
    """One column as it appears in the original workbook."""

    index: int  # 1-based
    group: str | None  # merged group title on the header row, if any
    label: str  # sub-header when grouped, else the header text
    width: float | None

    @property
    def key(self) -> str:
        return f"{self.group} | {self.label}" if self.group else self.label


@dataclass
class OriginalSheet:
    """Everything needed to reproduce a workbook's shape and compare its data.

    Usually read from a workbook ST exported (:func:`read_original`). It can
    also be **synthesised from the API alone** (:func:`synthesize`), which is
    what lets the tool build a selector nobody downloaded by hand. A
    synthesised sheet has ``path is None`` and no rows, so there is nothing to
    diff against and :attr:`is_synthetic` says so rather than the code
    inferring it from an empty ``parts``.
    """

    path: Path | None
    stem: str
    sheet_title: str
    breadcrumb: str
    columns: list[SheetColumn]
    parts: list[str]
    data: dict[str, dict[str, object]]  # part -> column key -> value
    has_sub_header: bool
    data_start_row: int
    logo_png: bytes | None = None
    logo_size: tuple[int, int] | None = None
    blank_rendering: str = "-"
    duplicate_parts: list[str] = field(default_factory=list)
    #: True when there was no source workbook -- the schema came from the API.
    is_synthetic: bool = False
    #: Set for a synthesised sheet, whose title is not in a breadcrumb.
    api_level_title: str | None = None

    @property
    def level_title(self) -> str:
        """Last breadcrumb segment -- the API's ``levelTitle`` for this file."""
        if self.api_level_title:
            return self.api_level_title
        return self.breadcrumb.rsplit("/", 1)[-1].strip() if self.breadcrumb else self.stem

    @property
    def column_keys(self) -> list[str]:
        return [c.key for c in self.columns]


def _group_titles(worksheet) -> dict[int, str]:
    """Map every column covered by a horizontal merge on the header row to
    that merge's title, so merged continuation cells stop reading as None."""
    groups: dict[int, str] = {}
    for rng in worksheet.merged_cells.ranges:
        if rng.min_row == HEADER_ROW == rng.max_row:
            title = worksheet.cell(HEADER_ROW, rng.min_col).value
            if title is not None:
                for col in range(rng.min_col, rng.max_col + 1):
                    groups[col] = str(title).strip()
    return groups


def _uses_sub_header(worksheet) -> bool:
    """Grouped files merge the ungrouped headers down over rows 10-11."""
    return any(
        rng.min_row == HEADER_ROW and rng.max_row == HEADER_ROW + 1
        for rng in worksheet.merged_cells.ranges
    )


def read_original(path: str | Path) -> OriginalSheet:
    path = Path(path)
    with warnings.catch_warnings():
        # ST's exports ship no default style; openpyxl warns and then does
        # the right thing. Nothing here depends on the default style.
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(path)
    worksheet = workbook.active

    has_sub = _uses_sub_header(worksheet)
    sub_row = HEADER_ROW + 1
    data_start = sub_row + 1 if has_sub else HEADER_ROW + 1
    groups = _group_titles(worksheet)

    columns: list[SheetColumn] = []
    for index in range(1, worksheet.max_column + 1):
        header = groups.get(index, worksheet.cell(HEADER_ROW, index).value)
        sub = worksheet.cell(sub_row, index).value if has_sub else None
        header = str(header).strip() if header is not None else ""
        sub = str(sub).strip() if sub is not None else None
        width = worksheet.column_dimensions[get_column_letter(index)].width
        if sub:
            columns.append(SheetColumn(index, header, sub, width))
        else:
            columns.append(SheetColumn(index, None, header, width))

    parts: list[str] = []
    seen: set[str] = set()
    duplicates: list[str] = []
    data: dict[str, dict[str, object]] = {}
    for row in range(data_start, worksheet.max_row + 1):
        part = worksheet.cell(row, 1).value
        if part is None or not str(part).strip():
            continue
        part = str(part).strip()
        if part in seen:
            duplicates.append(part)
            continue
        seen.add(part)
        parts.append(part)
        data[part] = {c.key: worksheet.cell(row, c.index).value for c in columns}

    logo, logo_size = _extract_logo(worksheet)

    breadcrumb = worksheet.cell(BREADCRUMB_ROW, 1).value
    return OriginalSheet(
        path=path,
        stem=path.stem.replace(" - Products", ""),
        sheet_title=worksheet.title,
        breadcrumb=str(breadcrumb).strip() if breadcrumb else "",
        columns=columns,
        parts=parts,
        data=data,
        has_sub_header=has_sub,
        data_start_row=data_start,
        logo_png=logo,
        logo_size=logo_size,
        duplicate_parts=duplicates,
    )


#: Widths for a synthesised sheet, which has no workbook to copy them from.
SYNTHETIC_WIDTH = 18.0
SYNTHETIC_WIDE_WIDTH = 44.0
_WIDE_COLUMNS = ("General Description", "Package", "Additional Interfaces")


def synthesize(grid, stem: str | None = None, *, logo_png: bytes | None = None,
               logo_size: tuple[int, int] | None = None) -> OriginalSheet:
    """Build a sheet skeleton from an API grid, with no workbook to copy.

    This is what makes a *discovered* selector buildable. ST composes its
    export headers from column metadata rather than returning them --
    ``name [" (symbol)"] [" (conditional)"] [" qualifier"]``, with an
    ``aggregation`` key becoming a merged group header -- and
    :attr:`stproducts.api.Column.key` already implements exactly that rule.
    So the schema ST would have exported is recoverable from the grid, and
    :func:`schema_matches` checks that claim against the nine workbooks that
    *were* exported, on every run.

    What cannot be recovered is the ST logo (a PNG embedded in the file, not
    in any API response) and the column widths. Widths get a sane default;
    the logo is passed in from a workbook that has one, or omitted --
    ``writer.py`` already guards on both.
    """
    columns: list[SheetColumn] = []
    for index, column in enumerate(grid.columns, start=1):
        wide = column.label in _WIDE_COLUMNS or column.name in _WIDE_COLUMNS
        columns.append(SheetColumn(
            index=index,
            group=column.aggregation or None,
            label=column.label,
            width=SYNTHETIC_WIDE_WIDTH if wide else SYNTHETIC_WIDTH,
        ))

    title = grid.level_title.strip()
    breadcrumb = getattr(grid, "breadcrumb", "") or f"Microcontrollers & microprocessors/{title}"
    has_sub = any(c.group for c in columns)
    return OriginalSheet(
        path=None,
        stem=stem or title,
        sheet_title="ProductsList",
        breadcrumb=breadcrumb,
        columns=columns,
        parts=[],          # nothing to compare against
        data={},
        has_sub_header=has_sub,
        data_start_row=HEADER_ROW + (2 if has_sub else 1),
        logo_png=logo_png,
        logo_size=logo_size,
        is_synthetic=True,
        api_level_title=title,
    )


def unreproducible_headers(sheet: OriginalSheet, grid) -> list[str]:
    """Shipped headers that :func:`synthesize` fails to reproduce from the API.

    Free, continuous evidence that the header-composition rule is sound. The
    nine workbooks in ``product_selector/`` are ground truth ST itself
    exported, so any header the rule cannot rebuild means a discovered
    selector would be built with a wrong schema. Currently zero across all
    330 shipped headers.

    The check runs one way only, and deliberately. A workbook is a snapshot
    of whatever columns ST served the day it was exported, and the grid has
    grown since -- the F2 file has 36 of today's 66. Columns present in the
    API and absent from the workbook are the ordinary appended-column case,
    not a failure. Columns in the workbook that the API cannot express are
    the real signal.
    """
    reproducible = set(synthesize(grid).column_keys)
    return [key for key in sheet.column_keys if key not in reproducible]


def _extract_logo(worksheet) -> tuple[bytes | None, tuple[int, int] | None]:
    """Pull the ST banner image out so the corrected file can carry it too."""
    for image in getattr(worksheet, "_images", []):
        try:
            data = image._data()
        except Exception:  # noqa: BLE001 -- a missing logo must not fail a run
            continue
        return data, (image.width, image.height)
    return None, None
